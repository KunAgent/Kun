#!/usr/bin/env python3
"""
将 Excel 配置清单解析为文本格式，用于传入 knot 智能体。

使用方式:
    python parse_excel.py --file /path/to/配置清单.xlsx
    python parse_excel.py --file /path/to/配置清单.xlsx --format markdown
    python parse_excel.py --file /path/to/配置清单.xlsx --format json

输出到 stdout，可直接管道传给 call_knot_agent.py。

依赖:
    pip install openpyxl
"""

import argparse
import json
import sys

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl。请执行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def _expand_merged_cells(ws) -> None:
    """把所有合并区域左上角的值复制到区域内每个 cell，并解除合并。

    openpyxl 行为：被合并区域里只有左上角 cell 持有真值，
    其余 cell 通过 iter_rows(values_only=True) 取到的都是 None。
    不补值会让下游 markdown 表里跨行/跨列合并的字段（站点/地域/计费模式/币种 等）
    在第二行起整列为空，远端 LLM 会因此误判或追问，违反铁律 1（忠实搬运）。

    in-place 修改临时 workbook 是安全的——load_workbook 不会写回原文件。
    """
    # 必须先 snapshot 成 list：unmerge_cells 会改 ws.merged_cells.ranges 集合
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).value = top_left_value


def _dedupe_headers(raw_headers: list[str]) -> list[str]:
    """表头行如果含水平合并（fill 后会出现同名列），加 _2/_3 后缀去重，
    避免 row_dict 用同一个 key 把多列数据互相覆盖（dict 丢列）。"""
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in raw_headers:
        if h not in seen:
            seen[h] = 1
            out.append(h)
        else:
            seen[h] += 1
            out.append(f"{h}_{seen[h]}")
    return out


def parse_excel(file_path: str, sheet_name: str = None) -> list[dict]:
    """
    解析 Excel 文件为结构化数据。

    Args:
        file_path: Excel 文件路径
        sheet_name: 指定 sheet 名称（默认取第一个）

    Returns:
        list[dict]: 每行数据作为一个字典，key 为表头列名
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    _expand_merged_cells(ws)

    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    # 第一行作为表头（合并展开后可能出现同名列，做去重保住所有数据列）
    raw_headers = [
        str(h).strip() if h is not None else f"列{i+1}"
        for i, h in enumerate(rows[0])
    ]
    headers = _dedupe_headers(raw_headers)

    # 解析数据行（跳过全空行）
    data = []
    for row in rows[1:]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(cell).strip() if cell is not None else ""
        data.append(row_dict)

    return data


def _sanitize_md_cell(value: str) -> str:
    """规整单元格内容，使其能安全放进 markdown 表格的「一行一格」。

    Excel 单元格可能含有内部换行（\\n / \\r）或竖线（|），二者都会破坏
    markdown 表格的行/列结构：换行会把一个逻辑商品行撑成多物理行（导致
    下游 split_tasks 按物理行 fan-out 时多切出残缺 task），竖线会错位列。

    处理（只搬运、不丢内容）：
    - 统一换行 \\r\\n / \\r → \\n，再把 \\n 压成单个空格
    - 竖线 | 转义为 \\| （markdown 表格内的字面竖线写法）
    - 合并连续空白为单空格，并去首尾空白
    """
    if not value:
        return ""
    # 1) 换行统一并压成空格（保留全部文字内容，仅去掉排版换行）
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = value.replace("\n", " ")
    # 2) 转义竖线，避免破坏列结构
    value = value.replace("|", "\\|")
    # 3) 合并连续空白
    value = " ".join(value.split())
    return value


def format_as_markdown(data: list[dict]) -> str:
    """将解析结果格式化为 Markdown 表格。"""
    if not data:
        return "（空表格）"

    headers = list(data[0].keys())

    # 构建表头（表头同样规整，避免表头含换行/竖线）
    lines = []
    lines.append("| " + " | ".join(_sanitize_md_cell(h) for h in headers) + " |")
    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

    # 构建数据行
    for row in data:
        cells = [_sanitize_md_cell(row.get(h, "")) for h in headers]
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def format_as_json(data: list[dict]) -> str:
    """将解析结果格式化为 JSON。"""
    return json.dumps(data, ensure_ascii=False, indent=2)


def format_as_text(data: list[dict]) -> str:
    """将解析结果格式化为逐行描述的纯文本。"""
    if not data:
        return "（空表格）"

    lines = []
    for i, row in enumerate(data, 1):
        parts = []
        for k, v in row.items():
            if v:  # 跳过空值
                parts.append(f"{k}: {v}")
        lines.append(f"第{i}行: {', '.join(parts)}")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="将 Excel 配置清单解析为文本格式")
    parser.add_argument("--file", "-f", required=True, help="Excel 文件路径")
    parser.add_argument("--sheet", "-s", default=None, help="Sheet 名称（默认取第一个）")
    parser.add_argument(
        "--format", choices=["markdown", "json", "text"], default="markdown",
        help="输出格式: markdown（默认）、json、text"
    )

    args = parser.parse_args()

    # 解析 Excel
    try:
        data = parse_excel(args.file, args.sheet)
    except FileNotFoundError:
        print(f"错误: 文件不存在: {args.file}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"错误: 解析 Excel 失败: {e}", file=sys.stderr)
        sys.exit(1)

    if not data:
        print("警告: 未解析到任何数据行", file=sys.stderr)
        sys.exit(0)

    # 格式化输出
    if args.format == "markdown":
        output = format_as_markdown(data)
    elif args.format == "json":
        output = format_as_json(data)
    else:
        output = format_as_text(data)

    print(output)


if __name__ == "__main__":
    main()
