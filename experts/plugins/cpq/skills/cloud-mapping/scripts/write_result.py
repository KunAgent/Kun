#!/usr/bin/env python3
"""
write_result.py

将 cloud-mapping 映射结果写入 Excel 文件。

用法:
  echo '<json>' | python3 write_result.py <output.xlsx>
  python3 write_result.py <output.xlsx> <input.json>

输入 JSON 格式（数组，每项 4 个字段）:
  [{ "原规格描述": "...", "腾讯云产品": "...", "腾讯云规格": "...", "备注": "..." }]

也接受英文 key 别名:
  source_summary -> 原规格描述
  target_product -> 腾讯云产品
  target_spec    -> 腾讯云规格
  notes          -> 备注
"""

import json
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = ["原规格描述", "腾讯云产品", "腾讯云规格", "备注"]
ALIAS = {
    "source_summary": "原规格描述",
    "target_product": "腾讯云产品",
    "target_spec": "腾讯云规格",
    "notes": "备注",
}


def normalize(row: dict) -> dict:
    out = {}
    for k, v in row.items():
        key = ALIAS.get(k, k)
        if key in HEADERS:
            out[key] = v or ""
    return out


import re

# 合法 provenance 前缀（必须出现在备注中）
_VALID_PROVENANCE_RE = re.compile(
    r"(?:dict:[a-zA-Z0-9_\-./]+|migraq\(|user(?::|$|\s)|unresolved|\[unresolved\])"
)


def validate_provenance(rows: list[dict]) -> list[str]:
    """
    校验规则：
    1. 如果某行有"腾讯云产品"或"腾讯云规格"的值，备注中必须包含合法 provenance
    2. 如果某行标了 [unresolved]，备注中必须提及 migraq（表明已尝试兜底）
    违反任一条均视为 error，拒绝写入。
    """
    errors = []
    for i, row in enumerate(rows, 1):
        product = (row.get("腾讯云产品") or "").strip()
        spec = (row.get("腾讯云规格") or "").strip()
        notes = row.get("备注") or ""

        # 规则1：有目标产品或规格时，必须有合法来源
        if product or spec:
            if not _VALID_PROVENANCE_RE.search(notes):
                errors.append(
                    f"  行 {i}: 原规格描述=\"{row.get('原规格描述', '')[:40]}\" "
                    f"有目标值但备注中缺少合法 provenance (dict:/migraq(/user/[unresolved])"
                )

        # 规则2：标了 [unresolved] 时，必须说明已尝试 migraq 兜底
        if "[unresolved]" in notes and "migraq" not in notes.lower():
            errors.append(
                f"  行 {i}: 原规格描述=\"{row.get('原规格描述', '')[:40]}\" "
                f"标了 [unresolved] 但未说明 migraq 兜底结果，必须先调用 migraq 再标 unresolved"
            )

    return errors


def write_xlsx(rows: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "映射结果"

    hfont = Font(bold=True)
    hfill = PatternFill("solid", fgColor="D9E1F2")
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 2):
        for j, h in enumerate(HEADERS, 1):
            ws.cell(row=i, column=j, value=row.get(h, ""))

    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 60)

    wb.save(output_path)
    print(f"wrote {len(rows)} rows to {output_path}")


def main():
    if len(sys.argv) < 2 or sys.argv[1] in ("--help", "-h"):
        print(__doc__.strip())
        sys.exit(0)

    output_path = sys.argv[1]

    if len(sys.argv) >= 3:
        with open(sys.argv[2], "r", encoding="utf-8") as f:
            raw = json.load(f)
    else:
        raw = json.loads(sys.stdin.read())

    if isinstance(raw, dict):
        raw = raw.get("mappings") or raw.get("data") or raw.get("items") or []

    if not isinstance(raw, list) or len(raw) == 0:
        print("错误: 输入 JSON 为空数组", file=sys.stderr)
        sys.exit(1)

    rows = [normalize(r) for r in raw]

    # 强制校验
    errors = validate_provenance(rows)
    if errors:
        print("❌ 校验失败，拒绝写入：", file=sys.stderr)
        print("\n".join(errors), file=sys.stderr)
        print("\n规则：", file=sys.stderr)
        print("  1. 有目标值 → 备注必须含 dict:<file> / migraq(session:<id>) / user", file=sys.stderr)
        print("  2. 标 [unresolved] → 备注必须提及 migraq（表明已尝试兜底）", file=sys.stderr)
        print("\n请先调用 migraq skill 兜底，再重新生成结果。", file=sys.stderr)
        sys.exit(1)

    write_xlsx(rows, output_path)


if __name__ == "__main__":
    main()
