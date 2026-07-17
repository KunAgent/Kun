#!/usr/bin/env python3
"""
write_result.py (cloud-mapping-intl variant)

将 cloud-mapping-intl 映射结果写入 Excel 文件。

与 cn 版 `plugins/cpq/skills/cloud-mapping/scripts/write_result.py` 的差异：
  - 删除「[unresolved] 必须提及 migraq」的强制规则，因为 intl 红线 #2 明令
    禁止调用 migraq，未命中直接 `[unresolved] dict_not_found:` 即可。
  - 其它行为（HEADERS / ALIAS / 字典来源校验 / Excel 排版）与 cn 版完全一致。

用法:
  echo '<json>' | python3 write_result.py <output.xlsx>
  python3 write_result.py <output.xlsx> <input.json>

输入 JSON 格式（数组，每项 4 个字段）:
  [{ "原规格描述": "...", "腾讯云产品": "...", "腾讯云规格": "...", "备注": "..." }]

也接受英文 key 别名:
  source_summary -> 原规格描述
  target_product -> 腾讯云产品
  target_spec    -> 腾讯云规格
  notes          -> 备注
"""

import json
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = ["原规格描述", "腾讯云产品", "腾讯云规格", "备注"]
ALIAS = {
    "source_summary": "原规格描述",
    "target_product": "腾讯云产品",
    "target_spec": "腾讯云规格",
    "notes": "备注",
}

# 合法 provenance 前缀（必须出现在备注中）
# 与 cn 版一致；intl 不会出现 migraq(，但保留兼容（cn 数据派生时不会带入）。
_VALID_PROVENANCE_RE = re.compile(
    r"(?:dict:[a-zA-Z0-9_\-./]+|user(?::|$|\s)|unresolved|\[unresolved\])"
)


def normalize(row: dict) -> dict:
    out = {}
    for k, v in row.items():
        key = ALIAS.get(k, k)
        if key in HEADERS:
            out[key] = v or ""
    return out


def validate_provenance(rows: list[dict]) -> list[str]:
    """
    intl 校验规则：
    1. 如果某行有"腾讯云产品"或"腾讯云规格"的值，备注中必须包含合法 provenance
       (dict:<file> / user / [unresolved])
    违反视为 error，拒绝写入。

    与 cn 版的差异：
    - 不再要求 [unresolved] 行必须出现 "migraq" 字样，因为 intl 红线 #2
      明令禁止 migraq 兜底；intl 未命中 = unresolved，无须 migraq 证据。
    """
    errors = []
    for i, row in enumerate(rows, 1):
        product = (row.get("腾讯云产品") or "").strip()
        spec = (row.get("腾讯云规格") or "").strip()
        notes = row.get("备注") or ""

        if product or spec:
            if not _VALID_PROVENANCE_RE.search(notes):
                errors.append(
                    f"  行 {i}: 原规格描述=\"{row.get('原规格描述', '')[:40]}\" "
                    f"有目标值但备注中缺少合法 provenance (dict:/user/[unresolved])"
                )

    return errors


def write_xlsx(rows: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "映射结果"

    hfont = Font(bold=True)
    hfill = PatternFill("solid", fgColor="D9E1F2")
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 2):
        for j, h in enumerate(HEADERS, 1):
            ws.cell(row=i, column=j, value=row.get(h, ""))

    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 60)

    wb.save(output_path)
    print(f"wrote {len(rows)} rows to {output_path}")


def main():
    if len(sys.argv) < 2 or sys.argv[1] in ("--help", "-h"):
        print(__doc__.strip())
        sys.exit(0)

    output_path = sys.argv[1]

    if len(sys.argv) >= 3:
        with open(sys.argv[2], "r", encoding="utf-8") as f:
            raw = json.load(f)
    else:
        raw = json.loads(sys.stdin.read())

    if isinstance(raw, dict):
        raw = raw.get("mappings") or raw.get("data") or raw.get("items") or []

    if not isinstance(raw, list) or len(raw) == 0:
        print("错误: 输入 JSON 为空数组", file=sys.stderr)
        sys.exit(1)

    rows = [normalize(r) for r in raw]

    errors = validate_provenance(rows)
    if errors:
        print("❌ 校验失败，拒绝写入：", file=sys.stderr)
        print("\n".join(errors), file=sys.stderr)
        print("\n规则：", file=sys.stderr)
        print("  1. 有目标值 → 备注必须含 dict:<file> / user / [unresolved]", file=sys.stderr)
        print(
            "\nintl 字典未命中直接标 `[unresolved] dict_not_found: ...`，"
            "无须 migraq 兜底（与 cn 版差异点）。",
            file=sys.stderr,
        )
        sys.exit(1)

    write_xlsx(rows, output_path)


if __name__ == "__main__":
    main()
