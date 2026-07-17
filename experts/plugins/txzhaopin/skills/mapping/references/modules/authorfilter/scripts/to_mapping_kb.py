#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
authorfilter → org-knowledge-base 转换器

读取 AuthorFilter 标注完成的 Excel（J/K/L 列）→ 按公司主体合并入 knowledge-base/org-mapping/{company}.json
+ 生成 HTML 架构图

用法：
    python3 to_mapping_kb.py \\
        --excel /path/to/cvpr2026.xlsx \\
        --workspace /path/to/workspace \\
        --venue "CVPR 2026" \\
        --source-note "标注于 2026-06-10"

依赖：openpyxl
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.cell.rich_text import CellRichText
except ImportError:
    print("[ERROR] 需要安装 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================
# 公司主体识别表（25 个标准化主体）
# ============================================================
COMPANY_ALIASES = {
    # 阿里系
    "alibaba": {
        "name": "阿里巴巴",
        "name_en": "Alibaba Group",
        "patterns": [r"阿里巴巴", r"\balibaba\b", r"达摩院", r"DAMO Academy", r"淘天", r"Taobao", r"Tmall", r"通义", r"Qwen"],
    },
    "ant-group": {
        "name": "蚂蚁集团",
        "name_en": "Ant Group",
        "patterns": [r"蚂蚁集团", r"\bant\s*group\b", r"\bant\s*financial\b", r"支付宝", r"Alipay"],
    },
    # 字节
    "bytedance": {
        "name": "字节跳动",
        "name_en": "ByteDance",
        "patterns": [r"字节", r"\bbytedance\b", r"\bbyte\s*dance\b", r"\btiktok\b", r"\bdouyin\b", r"抖音", r"\bseed\b.*byte", r"\bbyte.*seed\b"],
    },
    # 腾讯
    "tencent": {
        "name": "腾讯",
        "name_en": "Tencent",
        "patterns": [r"腾讯", r"\btencent\b", r"WeChat", r"微信", r"\bARC\s*Lab\b", r"优图", r"YouTu", r"Hunyuan", r"混元"],
    },
    # 华为
    "huawei": {
        "name": "华为",
        "name_en": "Huawei",
        "patterns": [r"华为", r"\bhuawei\b", r"诺亚方舟", r"Noah's Ark", r"昇腾", r"Ascend", r"\bHiSilicon\b"],
    },
    # 百度
    "baidu": {
        "name": "百度",
        "name_en": "Baidu",
        "patterns": [r"百度", r"\bbaidu\b", r"Apollo", r"文心", r"\bERNIE\b"],
    },
    # 商汤
    "sensetime": {
        "name": "商汤",
        "name_en": "SenseTime",
        "patterns": [r"商汤", r"\bsensetime\b", r"\bsensenova\b", r"日日新"],
    },
    # 快手
    "kuaishou": {
        "name": "快手",
        "name_en": "Kuaishou",
        "patterns": [r"快手", r"\bkuaishou\b"],
    },
    # 京东
    "jd": {
        "name": "京东",
        "name_en": "JD.COM",
        "patterns": [r"京东", r"\bjd\.com\b", r"\bjd\s*ai\b", r"JD Research"],
    },
    # 美团
    "meituan": {
        "name": "美团",
        "name_en": "Meituan",
        "patterns": [r"美团", r"\bmeituan\b"],
    },
    # 小米
    "xiaomi": {
        "name": "小米",
        "name_en": "Xiaomi",
        "patterns": [r"小米", r"\bxiaomi\b", r"小米汽车", r"Xiaomi EV"],
    },
    # 地平线
    "horizon-robotics": {
        "name": "地平线",
        "name_en": "Horizon Robotics",
        "patterns": [r"地平线", r"\bhorizon\s*robotics\b"],
    },
    # 上海AI Lab
    "shanghai-ai-lab": {
        "name": "上海AI Lab",
        "name_en": "Shanghai AI Laboratory",
        "patterns": [r"上海\s*AI\s*Lab", r"\bshanghai\s*ai\s*lab", r"\bshanghai\s*ai\s*laboratory\b", r"\bShlab\b"],
    },
    # 北京智源
    "baai": {
        "name": "北京智源",
        "name_en": "BAAI",
        "patterns": [r"智源", r"\bBAAI\b", r"Beijing Academy of Artificial Intelligence"],
    },
    # 鹏城实验室
    "pengcheng-lab": {
        "name": "鹏城实验室",
        "name_en": "Pengcheng Laboratory",
        "patterns": [r"鹏城", r"Pengcheng"],
    },
    # 智元机器人
    "agibot": {
        "name": "智元机器人",
        "name_en": "AGIBOT",
        "patterns": [r"智元", r"\bAGIBOT\b", r"\bAgibot\b"],
    },
    # 理想
    "lixiang": {
        "name": "理想汽车",
        "name_en": "Li Auto",
        "patterns": [r"理想汽车", r"\bli\s*auto\b", r"\bLiXiang\b", r"\bLi-Auto\b"],
    },
    # 小鹏
    "xpeng": {
        "name": "小鹏汽车",
        "name_en": "XPeng",
        "patterns": [r"小鹏", r"\bxpeng\b", r"\bXmotors\b"],
    },
    # 滴滴
    "didi": {
        "name": "滴滴",
        "name_en": "DiDi",
        "patterns": [r"滴滴", r"\bdidi\b"],
    },
    # ZTE 中兴
    "zte": {
        "name": "中兴",
        "name_en": "ZTE",
        "patterns": [r"中兴", r"\bZTE\b"],
    },
    # OPPO
    "oppo": {
        "name": "OPPO",
        "name_en": "OPPO",
        "patterns": [r"\bOPPO\b"],
    },
    # vivo
    "vivo": {
        "name": "vivo",
        "name_en": "vivo",
        "patterns": [r"\bvivo\b"],
    },
    # ----- 海外 -----
    "meta": {
        "name": "Meta",
        "name_en": "Meta Platforms",
        "patterns": [r"\bmeta\s*ai\b", r"\bmeta\s*reality\s*labs\b", r"\bmeta\s*FAIR\b", r"\bFAIR\b", r"\bmeta\s*platforms\b", r"^\s*meta\b"],
    },
    "google": {
        "name": "Google",
        "name_en": "Google",
        "patterns": [r"\bgoogle\b", r"\bdeepmind\b", r"\bgoogle\s*research\b"],
    },
    "microsoft": {
        "name": "微软",
        "name_en": "Microsoft",
        "patterns": [r"微软", r"\bmicrosoft\b", r"\bMSRA\b", r"\bMSR\s*Asia\b"],
    },
    "apple": {
        "name": "Apple",
        "name_en": "Apple Inc.",
        "patterns": [r"^\s*apple\b", r"\bapple\s*inc\b"],
    },
    "nvidia": {
        "name": "NVIDIA",
        "name_en": "NVIDIA",
        "patterns": [r"\bnvidia\b"],
    },
    "adobe": {
        "name": "Adobe",
        "name_en": "Adobe",
        "patterns": [r"\badobe\b"],
    },
    "amazon": {
        "name": "Amazon",
        "name_en": "Amazon",
        "patterns": [r"\bamazon\b", r"\bAWS\b"],
    },
    "openai": {
        "name": "OpenAI",
        "name_en": "OpenAI",
        "patterns": [r"\bopenai\b"],
    },
    "anthropic": {
        "name": "Anthropic",
        "name_en": "Anthropic",
        "patterns": [r"\banthropic\b"],
    },
}


# ============================================================
# 部门识别表（公司下的 AI Lab / Research Group）
# ============================================================
DEPT_ALIAS = {
    "tencent": [
        ("arc-lab", "ARC Lab", [r"\bARC\s*Lab\b"]),
        ("ai-lab", "AI Lab", [r"\btencent\s*ai\s*lab\b", r"腾讯\s*AI\s*Lab"]),
        ("hunyuan", "Hunyuan", [r"\bhunyuan\b", r"混元"]),
        ("wechat-vision", "WeChat Vision", [r"WeChat", r"微信"]),
        ("youtu", "优图实验室", [r"优图", r"\bYouTu\b"]),
    ],
    "alibaba": [
        ("damo", "达摩院", [r"达摩院", r"\bDAMO\b"]),
        ("tongyi", "通义实验室", [r"通义", r"\bQwen\b", r"Tongyi"]),
        ("taobao-tmall", "淘天集团", [r"淘天", r"Taobao", r"Tmall"]),
    ],
    "bytedance": [
        ("seed", "Seed Team", [r"\bSeed\b"]),
        ("research", "ByteDance Research", [r"Research"]),
    ],
    "huawei": [
        ("noah", "诺亚方舟实验室", [r"诺亚方舟", r"Noah'?s Ark", r"\bNoah\b"]),
    ],
    "baidu": [
        ("apollo", "Apollo", [r"Apollo"]),
        ("ernie", "文心团队", [r"文心", r"\bERNIE\b"]),
        ("research", "百度研究院", [r"研究院"]),
    ],
    "meta": [
        ("fair", "FAIR", [r"\bFAIR\b"]),
        ("reality-labs", "Reality Labs", [r"Reality Labs"]),
        ("meta-ai", "Meta AI", [r"Meta\s*AI"]),
    ],
    "google": [
        ("deepmind", "DeepMind", [r"DeepMind"]),
        ("research", "Google Research", [r"Google\s*Research"]),
        ("brain", "Google Brain", [r"Google\s*Brain"]),
    ],
    "microsoft": [
        ("msra", "MSR Asia", [r"\bMSRA\b", r"MSR\s*Asia"]),
        ("msr", "Microsoft Research", [r"Microsoft\s*Research"]),
    ],
}


# ============================================================
# 工具函数
# ============================================================
def now_iso():
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S+08:00")


def slugify(text):
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text, flags=re.UNICODE)
    text = re.sub(r"[\s_]+", "-", text)
    return text[:50]


def identify_company(l_value):
    """从 L 列值识别 company_id + 部门"""
    if not l_value or not l_value.strip():
        return None, None, None
    l = l_value.strip()
    for cid, info in COMPANY_ALIASES.items():
        for pat in info["patterns"]:
            if re.search(pat, l, re.IGNORECASE):
                team_id, team_name = identify_team(cid, l)
                return cid, team_id, team_name
    # 不在标准表中，作为独立公司
    return slugify(l), None, None


def identify_team(company_id, l_value):
    """从 L 列值识别部门"""
    if company_id not in DEPT_ALIAS:
        return None, None
    for tid, tname, pats in DEPT_ALIAS[company_id]:
        for pat in pats:
            if re.search(pat, l_value, re.IGNORECASE):
                return tid, tname
    return None, None


def normalize_name(name):
    """姓名标准化（用于去重）"""
    if not name:
        return ""
    return re.sub(r"\s+", " ", name.strip().lower())


def cell_text(cell):
    """提取 cell 文本（兼容 RichText）"""
    if cell is None or cell.value is None:
        return ""
    val = cell.value
    if isinstance(val, CellRichText):
        return "".join(str(t) for t in val)
    return str(val).strip()


def extract_bold_authors(cell):
    """从 F 列 RichText 提取加粗的作者名（企业作者）"""
    if cell is None or cell.value is None:
        return []
    val = cell.value
    if not isinstance(val, CellRichText):
        return []
    bold = []
    for block in val:
        try:
            font = getattr(block, "font", None)
            if font and getattr(font, "b", False):
                txt = str(block).strip(" ,;\t\n")
                if txt:
                    bold.append(txt)
        except Exception:
            continue
    return bold


def parse_authors_string(authors_str):
    """从 F 列文本切分作者列表"""
    if not authors_str:
        return []
    # 常见分隔符：逗号、分号、and
    raw = re.split(r"[,;]|\sand\s", authors_str)
    return [a.strip() for a in raw if a.strip()]


# ============================================================
# JSON 模板
# ============================================================
def empty_company_json(company_id):
    info = COMPANY_ALIASES.get(company_id, {})
    return {
        "company_id": company_id,
        "name": info.get("name", company_id),
        "name_en": info.get("name_en", ""),
        "industry": "Tech / AI / Internet",
        "industry_subcategory": "AI Research",
        "version": "1.0",
        "schema_version": "v4",
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "scope_note": "由 authorfilter v1.0 从顶会论文沉淀的 AI 研究员；可与其他 Mapping Skill 数据共存",
        "org_structure": {
            "id": "root",
            "name": info.get("name", company_id),
            "children": [
                {
                    "id": "dept-research",
                    "name": "AI 研究体系",
                    "children": []
                }
            ]
        },
        "personnel": [],
        "papers": [],
        "update_history": []
    }


def ensure_team(json_obj, team_id, team_name):
    """确保 org_structure 中有该 team"""
    if not team_id:
        return
    research = json_obj["org_structure"]["children"][0]
    for t in research["children"]:
        if t["id"] == team_id:
            return
    research["children"].append({"id": team_id, "name": team_name})


# ============================================================
# 核心：合并人员到公司 JSON
# ============================================================
def merge_person(json_obj, name, paper_info, team_id, team_name, l_raw, source_excel):
    """添加或更新一个企业作者"""
    # 找已有人
    existing = None
    for p in json_obj["personnel"]:
        if normalize_name(p.get("name", "")) == normalize_name(name):
            existing = p
            break

    if existing:
        # merge paper_history
        existing["paper_history"].append(paper_info)
        existing["updated_at"] = now_iso()
        existing["background_brief"] = f"{l_raw} · 已发表 {len(existing['paper_history'])} 篇顶会论文"
    else:
        person_id = f"person-{slugify(name)}-{json_obj['company_id']}"
        new_person = {
            "id": person_id[:80],
            "name": name,
            "affiliation_in_paper": l_raw,
            "department_id": "dept-research",
            "team_id": team_id,
            "team_name": team_name,
            "title": "AI Researcher",
            "title_abbr": "Researcher",
            "background_brief": f"{l_raw} · 已发表 1 篇顶会论文",
            "paper_history": [paper_info],
            "is_intern_likely": False,
            "source": "authorfilter v1.0",
            "source_urls": [],
            "confidence": "very_high",
            "added_at": now_iso(),
            "updated_at": now_iso(),
            "notes": f"论文首页明确标注 {l_raw}（来自 {source_excel}）"
        }
        json_obj["personnel"].append(new_person)


def merge_paper(json_obj, paper_record):
    """添加论文到公司 papers 列表（去重）"""
    pid = paper_record["paper_id"]
    for p in json_obj["papers"]:
        if p["paper_id"] == pid:
            return  # 已存在，跳过
    json_obj["papers"].append(paper_record)


# ============================================================
# HTML 渲染（极简版）
# ============================================================
def render_html(json_obj, out_path):
    info = COMPANY_ALIASES.get(json_obj["company_id"], {})
    name = json_obj["name"]
    name_en = json_obj.get("name_en", "")

    # 按部门分组人员
    by_team = {}
    for p in json_obj["personnel"]:
        t = p.get("team_id") or "_unknown"
        by_team.setdefault(t, []).append(p)

    # 生成 team 节点
    team_blocks = []
    for team in json_obj["org_structure"]["children"][0]["children"]:
        tid = team["id"]
        tname = team["name"]
        members = by_team.get(tid, [])
        if not members:
            continue
        person_html = ""
        for p in members:
            paper_count = len(p.get("paper_history", []))
            person_html += f'<div class="person"><b>{p["name"]}</b> <span class="papers">📄 {paper_count}</span></div>'
        team_blocks.append(f'''<div class="team">
  <div class="team-name">{tname} <span class="count">({len(members)} 人)</span></div>
  {person_html}
</div>''')

    # 没归类到 team 的（_unknown）
    unknown = by_team.get("_unknown", [])
    if unknown:
        person_html = ""
        for p in unknown:
            paper_count = len(p.get("paper_history", []))
            person_html += f'<div class="person"><b>{p["name"]}</b> <span class="papers">📄 {paper_count}</span></div>'
        team_blocks.append(f'''<div class="team">
  <div class="team-name">未归类部门 <span class="count">({len(unknown)} 人)</span></div>
  {person_html}
</div>''')

    teams_html = "\n".join(team_blocks) if team_blocks else "<div class='empty'>暂无入库人员</div>"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>{name} AI 研究体系 - 顶会论文沉淀</title>
<style>
body {{ font-family: -apple-system, "PingFang SC", "Microsoft YaHei", sans-serif; max-width: 1100px; margin: 30px auto; padding: 20px; background: #f5f5f7; }}
h1 {{ color: #1d1d1f; border-bottom: 3px solid #0071e3; padding-bottom: 10px; }}
.meta {{ color: #6e6e73; font-size: 14px; margin-bottom: 20px; }}
.dept {{ background: white; border-radius: 12px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.05); }}
.dept-title {{ font-size: 18px; font-weight: 600; color: #0071e3; margin-bottom: 16px; }}
.team {{ border-left: 4px solid #0071e3; padding: 12px 16px; margin: 12px 0; background: #f9f9fb; border-radius: 6px; }}
.team-name {{ font-weight: 600; font-size: 15px; color: #1d1d1f; margin-bottom: 8px; }}
.count {{ color: #6e6e73; font-weight: normal; font-size: 13px; }}
.person {{ display: inline-block; padding: 4px 10px; margin: 3px; background: white; border: 1px solid #d2d2d7; border-radius: 4px; font-size: 13px; }}
.papers {{ color: #0071e3; font-size: 12px; margin-left: 4px; }}
.empty {{ color: #6e6e73; font-style: italic; padding: 20px; text-align: center; }}
.footer {{ margin-top: 30px; padding: 16px; background: #fff; border-radius: 8px; font-size: 13px; color: #6e6e73; }}
</style>
</head>
<body>
<h1>{name} <span style="font-weight: normal; font-size: 20px; color: #6e6e73;">{name_en}</span></h1>
<div class="meta">
  AI 研究体系组织图（来自顶会论文公开 affiliation 沉淀）<br>
  最后更新：{json_obj["updated_at"]} ·
  入库 <b>{len(json_obj["personnel"])}</b> 位企业作者 ·
  覆盖 <b>{len(json_obj["papers"])}</b> 篇论文 ·
  来源：authorfilter v1.0
</div>

<div class="dept">
  <div class="dept-title">AI 研究体系</div>
  {teams_html}
</div>

<div class="footer">
  <b>数据来源</b>：本图所有人员均来自顶会论文（CVPR / ICCV / NeurIPS / ICML 等）首页 affiliation 公开标注。
  📄 数字代表该人员入库的论文数量。<br>
  <b>验证建议</b>：如需确认人员当前是否仍在该公司，可调用 linkedin-deep-miner 做时效性验证。
</div>
</body>
</html>"""

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


# ============================================================
# 主流程
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="AuthorFilter Excel → org-knowledge-base 转换器")
    parser.add_argument("--excel", required=True, help="标注完成的 Excel 文件")
    parser.add_argument("--workspace", required=True, help="工作区根目录（含 knowledge-base/）")
    parser.add_argument("--venue", required=True, help="顶会名称，如 CVPR 2026")
    parser.add_argument("--source-note", default="", help="备注（写入 update_history）")
    parser.add_argument("--title-col", type=int, default=5, help="标题列（默认 E=5）")
    parser.add_argument("--authors-col", type=int, default=6, help="作者列（默认 F=6）")
    parser.add_argument("--direction-col", type=int, default=10, help="方向列（默认 J=10）")
    parser.add_argument("--identity-col", type=int, default=11, help="身份列（默认 K=11）")
    parser.add_argument("--org-col", type=int, default=12, help="组织列（默认 L=12）")
    args = parser.parse_args()

    excel_path = args.excel
    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel 文件不存在：{excel_path}", file=sys.stderr)
        sys.exit(1)

    workspace = args.workspace
    kb_dir = Path(workspace) / "knowledge-base" / "org-mapping"
    charts_dir = kb_dir / "charts"
    kb_dir.mkdir(parents=True, exist_ok=True)
    charts_dir.mkdir(parents=True, exist_ok=True)

    print(f"[OPEN] {excel_path}")
    wb = openpyxl.load_workbook(excel_path, rich_text=True)
    ws = wb.active
    excel_basename = os.path.basename(excel_path)

    # 收集 → 按公司分组
    company_data = {}  # {company_id: json_obj}
    total_rows = 0
    industry_rows = 0
    skipped_no_l = 0
    venue_year_match = re.search(r"(\d{4})", args.venue)
    venue_year = int(venue_year_match.group(1)) if venue_year_match else None

    for row in range(2, ws.max_row + 1):
        title = cell_text(ws.cell(row=row, column=args.title_col))
        if not title:
            continue
        total_rows += 1

        identity = cell_text(ws.cell(row=row, column=args.identity_col))
        if identity != "工业界":
            continue
        industry_rows += 1

        l_raw = cell_text(ws.cell(row=row, column=args.org_col))
        if not l_raw:
            skipped_no_l += 1
            continue

        direction = cell_text(ws.cell(row=row, column=args.direction_col))
        authors_str = cell_text(ws.cell(row=row, column=args.authors_col))

        # 提取 F 列加粗作者（企业作者）
        bold_authors = extract_bold_authors(ws.cell(row=row, column=args.authors_col))
        if not bold_authors:
            # fallback: 没有加粗信息，跳过（让用户回去补加粗）
            print(f"[WARN] Row {row} K=工业界但 F 列没有加粗作者，跳过：{title[:60]}")
            continue

        # L 列可能含多个公司（用 " / " 或 "+" 分隔），逐个处理
        l_parts = re.split(r"\s*[/+]\s*|\s*、\s*", l_raw)

        # 该论文的所有公司+作者归属
        for l_part in l_parts:
            if not l_part.strip():
                continue
            company_id, team_id, team_name = identify_company(l_part)
            if not company_id:
                continue

            # 初始化公司 JSON
            if company_id not in company_data:
                # 优先读取已有
                existing_path = kb_dir / f"{company_id}.json"
                if existing_path.exists():
                    try:
                        with open(existing_path, "r", encoding="utf-8") as f:
                            company_data[company_id] = json.load(f)
                        # 确保新增字段存在
                        company_data[company_id].setdefault("papers", [])
                        company_data[company_id].setdefault("personnel", [])
                        company_data[company_id].setdefault("update_history", [])
                    except Exception as e:
                        print(f"[WARN] 已有 JSON 损坏，新建：{e}")
                        company_data[company_id] = empty_company_json(company_id)
                else:
                    company_data[company_id] = empty_company_json(company_id)

            j_obj = company_data[company_id]
            ensure_team(j_obj, team_id, team_name)

            # 构造论文记录
            paper_id = f"paper-{slugify(args.venue)}-{slugify(title[:40])}"
            paper_record = {
                "paper_id": paper_id,
                "title": title,
                "venue": args.venue,
                "venue_year": venue_year,
                "research_direction": direction,
                "company_authors": [
                    {"name": a, "team": team_name or ""} for a in bold_authors
                ],
                "first_author": parse_authors_string(authors_str)[0] if parse_authors_string(authors_str) else "",
                "added_at": now_iso(),
                "source_excel": f"{excel_basename} Row {row}"
            }

            # 把每个加粗作者加入该公司
            paper_info_for_person = {
                "paper_id": paper_id,
                "paper_title": title,
                "venue": args.venue,
                "research_direction": direction,
                "co_authors_at_company": [a for a in bold_authors],
            }
            for a in bold_authors:
                merge_person(j_obj, a, paper_info_for_person, team_id, team_name, l_part.strip(), excel_basename)

            merge_paper(j_obj, paper_record)

    # 写出每个公司 JSON + HTML
    written = []
    for cid, j_obj in company_data.items():
        j_obj["updated_at"] = now_iso()
        # 添加本次 update history
        j_obj["update_history"].append({
            "timestamp": now_iso(),
            "source": "authorfilter v1.0",
            "changes": f"从 {excel_basename} 入库 {len(j_obj['personnel'])} 位企业作者 / {len(j_obj['papers'])} 篇论文（{args.venue}）。{args.source_note}"
        })

        json_path = kb_dir / f"{cid}.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(j_obj, f, ensure_ascii=False, indent=2)
        # 渲染 HTML
        html_path = charts_dir / f"{cid}.html"
        render_html(j_obj, str(html_path))
        written.append((cid, len(j_obj["personnel"]), len(j_obj["papers"]), str(json_path), str(html_path)))

    # 输出报告
    print()
    print("=" * 70)
    print("[报告] AuthorFilter → org-knowledge-base 入库完成")
    print("=" * 70)
    print(f"Excel: {excel_basename}")
    print(f"会议: {args.venue}")
    print(f"扫描行数: {total_rows} | 工业界论文: {industry_rows} | 跳过(L列空): {skipped_no_l}")
    print(f"涉及公司: {len(written)}")
    print()
    for cid, p_count, paper_count, jp, hp in sorted(written, key=lambda x: -x[1]):
        info = COMPANY_ALIASES.get(cid, {})
        cn_name = info.get("name", cid)
        print(f"  {cn_name:20s}  人员: {p_count:3d}  论文: {paper_count:3d}")
        print(f"    JSON: {jp}")
        print(f"    HTML: {hp}")
    print()


if __name__ == "__main__":
    main()
