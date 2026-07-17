#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
面试数据清洗脚本（Skill1 场景简化版）
=====================================

用途：interview-data-processor Skill 的一键清洗
输入：原始 Excel/CSV 文件
输出：cleaned_data.json（标准化记录）+ quality_report.md（质量报告）

使用方式：
  python3 data_clean.py --input interview_data.xlsx
  python3 data_clean.py --input data.csv --output cleaned.json --report report.md
"""

import json
import re
import sys
import os
from pathlib import Path
from collections import defaultdict
from datetime import datetime

# ============================================================================
# 配置
# ============================================================================

MIN_COMMENT_LENGTH = 10

# 字段映射：标准字段 → 可能的原始列名（按优先级排序）
FIELD_MAPPING = {
    'name': ['候选人ID', '简历ID', '候选人', '姓名', '名字', 'name', '候选人名称', '员工姓名', 'candidate_id'],
    'comment': ['面试评价', '文字评价', '评价', '面试评语', '评语', 'comment', '综合评价', '考核评价', '能力评价'],
    'grade': ['面试评级', '面试结果', '面试等级', '评价等级', '评级', '等级', 'grade', 'rating', '绩效等级'],
    'result': ['面试流程结果', '本环节处理意见', '面试流程', '是否通过', '结果', 'result', '录用结果', '考核结论'],
    'round': ['环节名称', '面试环节', '面试轮次', '轮次', 'round', '面试阶段'],
    'position': ['职位名称（岗位）', '职位名称(岗位)', '面试职位', 'position', '职位名称'],
    'bg': ['招聘岗位BG', '面试BG', '事业群', 'BG', 'bg', 'department'],
    'dept': ['招聘岗位组织', '面试组织全路径', '部门', '事业部', '组织', '所属部门'],
    'candidateLevel': ['建议职级', '候选人职级', '职级', 'level', '建议级别'],
    'date': ['流程处理时间', '本环节处理时间', '面试时间', '面试日期', 'date'],
    'recruitPosition': ['招聘岗位', '学生简历投递职位', '应聘岗位'],
    'positionFamily': ['招聘岗位职位族', '面试职位族', '职位族'],
    'positionCategory': ['招聘岗位职位类', '面试职位类', '职位类'],
    'source': ['数据来源', '来源', '招聘类型', 'source'],
}

# 映射优先级
FIELD_PRIORITY = ['name', 'comment', 'grade', 'result', 'round', 'position', 'bg', 'dept',
                  'candidateLevel', 'date', 'source', 'recruitPosition', 'positionFamily', 'positionCategory']

# 评级归一化
GRADE_MAP = {
    'S': ['S', 's', 'S级', '优秀', 'excellent', '5', '5分', '卓越', '超出期望', '远超期望'],
    'A+': ['A+', 'a+', 'A+级', '杰出'],
    'A': ['A', 'a', 'A级', '良好', 'good', '4', '4分', '符合期望', '达到期望', '优良'],
    'A-': ['A-', 'a-', 'A-级'],
    'B': ['B', 'b', 'B级', '一般', 'average', '3', '3分', '中等', '基本符合', '待提升'],
}

# 结果归一化
RESULT_MAP = {
    'pass': ['通过', '是', '录用', 'pass', 'yes', 'offer', '推荐录用', '达标', '合格', '提交'],
    'fail': ['淘汰', '否', '不通过', 'fail', 'no', 'reject', '不达标', '不合格', 'PIP', '放弃'],
    'pending': ['待定', '观望', 'pending', '需讨论', '待确认', '待观察', '需改进', '待提升'],
}

# BG 标准列表
BG_LIST = ['CDG', 'CSIG', 'IEG', 'PCG', 'TEG', 'WXG', 'S1', 'S2', 'S3']

# BG 全称 → 标准码 映射（字典，易维护）
BG_FULLNAME_MAP = {
    '互动娱乐': 'IEG', 'IEG': 'IEG',
    '云与智慧产业': 'CSIG', 'CSIG': 'CSIG',
    '企业发展': 'CDG', 'CDG': 'CDG',
    '平台与内容': 'PCG', 'PCG': 'PCG',
    '技术工程': 'TEG', 'TEG': 'TEG',
    '微信': 'WXG', 'WXG': 'WXG',
    '社交网络': 'S1', 'S1': 'S1',
    'S2': 'S2',
    'S3': 'S3',
}

# 数据源检测特征词
SOURCE_KEYWORDS = {
    '社招面评': ['候选人', '面试', '录用', '淘汰', '面试流程结果', '环节名称', '建议职级'],
    '校招面评': ['简历ID', '学生简历投递职位', '面试环节', '本环节处理意见', '校招', '学校'],
    '活水面评': ['活水', '内部', '转岗', '原部门'],
}

# ============================================================================
# 工具函数
# ============================================================================

def normalize_grade(grade):
    if not grade:
        return ''
    g = str(grade).strip()
    for norm, aliases in GRADE_MAP.items():
        if g.lower() in [a.lower() for a in aliases]:
            return norm
    return ''

def normalize_result(result):
    if not result:
        return ''
    r = str(result).strip().lower()
    # 优先精确匹配
    for norm, aliases in RESULT_MAP.items():
        if r in [a.lower() for a in aliases]:
            return norm
    # 回退到包含匹配（处理"通过（建议xxx）"等带额外内容的情况）
    for norm, aliases in RESULT_MAP.items():
        for a in aliases:
            if a.lower() in r or r in a.lower():
                return norm
    return ''

def normalize_bg(bg):
    if not bg:
        return ''
    upper = str(bg).strip().upper()
    for std in BG_LIST:
        if std in upper:
            return std
    # 全称匹配（字典查找）
    for name, code in BG_FULLNAME_MAP.items():
        if name in str(bg):
            return code
    return str(bg).strip()

def normalize_dept(dept, bg_name=''):
    if not dept:
        return ''
    trimmed = str(dept).strip()
    parts = re.split(r'[/／]', trimmed)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) <= 1:
        return trimmed
    first_upper = parts[0].upper()
    is_bg = False
    if bg_name and bg_name.upper() in first_upper:
        is_bg = True
    if not is_bg:
        is_bg = any(std in first_upper for std in BG_LIST)
    if not is_bg:
        is_bg = any(name in parts[0] for name in BG_FULLNAME_MAP)
    if is_bg:
        return parts[1]
    return trimmed

def parse_level(level_str):
    if not level_str:
        return None
    match = re.search(r'(\d+)', str(level_str))
    return int(match.group(1)) if match else None

def format_date(value):
    if not value:
        return ''
    s = str(value).strip()
    m = re.match(r'(\d{4})[/\-年](\d{1,2})[/\-月](\d{1,2})', s)
    if m:
        return f'{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}'
    try:
        d = datetime.fromisoformat(s.replace('/', '-').split(' ')[0])
        return d.strftime('%Y-%m-%d')
    except:
        return s

def detect_source(headers):
    header_str = '|'.join(headers).lower()
    best = ('社招面评', 0)
    for src_type, keywords in SOURCE_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw.lower() in header_str)
        if score > best[1]:
            best = (src_type, score)
    return best[0]

def infer_recruit_category(source_type, raw_source=''):
    rs = raw_source.lower() if raw_source else ''
    if '校招' in rs:
        return '校招'
    if '社招' in rs or '活水' in rs:
        return '社招'
    if source_type in ('校招面评',):
        return '校招'
    return '社招'

def is_level_in_social_range(level, recruit_category):
    if '社招' not in (recruit_category or ''):
        return True
    if level is None:
        return True
    return 7 <= level <= 11

# ============================================================================
# 主流程
# ============================================================================

def map_fields(headers):
    """字段映射：返回 { 标准字段: 列索引 }"""
    mapping = {}
    used = set()
    for field in FIELD_PRIORITY:
        aliases = FIELD_MAPPING.get(field, [])
        # 精确匹配
        idx = -1
        for i, h in enumerate(headers):
            if i in used:
                continue
            if h.lower().strip() in [a.lower() for a in aliases]:
                idx = i
                break
        # 包含匹配
        if idx < 0:
            for i, h in enumerate(headers):
                if i in used:
                    continue
                if any(a.lower() in h.lower() for a in aliases):
                    idx = i
                    break
        if idx >= 0:
            used.add(idx)
        mapping[field] = idx
    return mapping

def process_file(input_path):
    """处理 Excel/CSV 文件，返回 (records, metadata)"""
    ext = Path(input_path).suffix.lower()
    
    if ext == '.csv':
        import csv
        with open(input_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = list(reader)
    elif ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
            all_rows = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                sheet_rows = []
                for row in ws.iter_rows(values_only=True):
                    sheet_rows.append([str(c) if c is not None else '' for c in row])
                if len(sheet_rows) >= 2:
                    all_rows = sheet_rows
                    break  # 用第一个有数据的 sheet
            rows = all_rows
            wb.close()
        except ImportError:
            print("❌ 需要安装 openpyxl: pip install openpyxl", file=sys.stderr)
            sys.exit(1)
    else:
        print(f"❌ 不支持的格式: {ext}", file=sys.stderr)
        sys.exit(1)
    
    # 裁剪尾部空行（Excel 的"已使用范围"可能远大于实际数据行数）
    while len(rows) > 1 and all(not str(c).strip() for c in rows[-1]):
        rows.pop()
    
    if len(rows) < 2:
        print("❌ 文件中没有数据", file=sys.stderr)
        sys.exit(1)
    
    headers = [str(h).strip() for h in rows[0]]
    source_type = detect_source(headers)
    field_map = map_fields(headers)
    
    # 统计
    stats = {
        'total_raw': len(rows) - 1,
        'empty_rows': 0,
        'short_comment': 0,
        'level_filtered': 0,
        'has_huoshui': 0,
        'unmapped_fields': [f for f in FIELD_PRIORITY[:8] if field_map.get(f, -1) < 0],
    }
    
    records = []
    for i in range(1, len(rows)):
        row = rows[i]
        if not row or all(not str(c).strip() for c in row):
            stats['empty_rows'] += 1
            continue
        
        def get(field):
            idx = field_map.get(field, -1)
            if idx < 0 or idx >= len(row):
                return ''
            return str(row[idx]).strip() if row[idx] else ''
        
        # 基础字段
        raw_source = get('source') or source_type
        recruit_cat = infer_recruit_category(source_type, raw_source)
        
        # 活水检测（多字段联合检测，避免单字段为空时漏判）
        huoshui_keywords = ['活水', '内部转岗', '转岗', '内部流动']
        huoshi_fields = [raw_source, get('position'), get('recruitPosition')]
        is_huoshui = any(
            any(kw in f for kw in huoshui_keywords)
            for f in huoshi_fields if f
        )
        if is_huoshui:
            stats['has_huoshui'] += 1
            continue  # 跳过活水数据
        
        comment = get('comment')
        if not comment or len(comment) < MIN_COMMENT_LENGTH:
            stats['short_comment'] += 1
            continue
        
        # BG 和 dept 归一化
        raw_bg = get('bg')
        bg = normalize_bg(raw_bg)
        raw_dept = get('dept')
        dept = normalize_dept(raw_dept, bg)
        
        # 职级
        level_str = get('candidateLevel')
        level = parse_level(level_str)
        
        # 社招职级过滤
        if not is_level_in_social_range(level, recruit_cat):
            stats['level_filtered'] += 1
            continue
        
        record = {
            'name': get('name'),
            'comment': comment,
            'grade': get('grade'),
            'gradeNorm': normalize_grade(get('grade')),
            'result': get('result'),
            'resultNorm': normalize_result(get('result')),
            'round': get('round'),
            'position': get('position'),
            'bg': bg,
            'dept': dept,
            'candidateLevel': level_str,
            'date': format_date(get('date')),
            'recruitPosition': get('recruitPosition'),
            'positionFamily': get('positionFamily'),
            'positionCategory': get('positionCategory'),
            'sourceType': source_type,
            'recruitCategory': recruit_cat,
        }
        records.append(record)
    
    # 统计汇总
    stats['valid'] = len(records)
    stats['pass_count'] = sum(1 for r in records if r['resultNorm'] == 'pass')
    stats['fail_count'] = sum(1 for r in records if r['resultNorm'] == 'fail')
    stats['pass_rate'] = round(stats['pass_count'] / max(1, stats['valid']) * 100, 1)
    
    # 分布统计
    position_dist = defaultdict(int)
    bg_dist = defaultdict(int)
    dept_dist = defaultdict(int)
    grade_dist = defaultdict(int)
    recruit_dist = defaultdict(int)
    level_dist = {'10-11级': 0, '7-9级': 0, '无职级': 0}
    # dept × recruitCategory 交叉统计（用于报告表格）
    cross_dept_recruit = defaultdict(lambda: {'count': 0, 'pass_count': 0, 'grade_dist': defaultdict(int)})
    
    for r in records:
        position_dist[r['position'] or '未指定'] += 1
        if r['bg']:
            bg_dist[r['bg']] += 1
        if r['dept']:
            dept_dist[r['dept']] += 1
        if r['gradeNorm']:
            grade_dist[r['gradeNorm']] += 1
        recruit_dist[r['recruitCategory']] += 1
        # 交叉统计
        dk = r['dept'] or '未指定'
        rk = r['recruitCategory']
        cross_dept_recruit[(dk, rk)]['count'] += 1
        if r['resultNorm'] == 'pass':
            cross_dept_recruit[(dk, rk)]['pass_count'] += 1
        if r['gradeNorm']:
            cross_dept_recruit[(dk, rk)]['grade_dist'][r['gradeNorm']] += 1
        lv = parse_level(r['candidateLevel'])
        if lv is None:
            level_dist['无职级'] += 1
        elif lv >= 10:
            level_dist['10-11级'] += 1
        else:
            level_dist['7-9级'] += 1
    
    stats['position_dist'] = dict(position_dist)
    stats['bg_dist'] = dict(bg_dist)
    stats['dept_dist'] = dict(sorted(dept_dist.items(), key=lambda x: -x[1]))
    stats['grade_dist'] = dict(grade_dist)
    stats['recruit_dist'] = dict(recruit_dist)
    stats['level_dist'] = level_dist
    stats['cross_dept_recruit'] = {f"{k[0]}|{k[1]}": dict(v) for k, v in cross_dept_recruit.items()}
    stats['field_map'] = {k: headers[v] if v >= 0 else '未映射' for k, v in field_map.items()}
    
    return records, stats

def quality_check(records, stats):
    """质量检查，返回 (blockers, warnings)"""
    blockers = []
    warnings = []
    
    # 阻断
    if stats['valid'] < 60:
        blockers.append(f"有效数据不足 60 条（当前 {stats['valid']} 条），样本量太小无法建模。")
    for field in ['comment', 'result', 'grade', 'round', 'position', 'bg']:
        if field in stats['unmapped_fields']:
            blockers.append(f"必选字段 {field} 未在数据中找到对应列。")
    
    # 警告
    if stats['has_huoshui'] > 0:
        warnings.append(f"数据中包含活水数据 {stats['has_huoshui']} 条，已自动排除。当前仅支持社招和校招。")
    if 60 <= stats['valid'] < 100:
        warnings.append(f"有效数据仅 {stats['valid']} 条，建模结果置信度较低，建议 100 条以上。")
    if stats['pass_rate'] < 10 or stats['pass_rate'] > 90:
        warnings.append(f"通过样本占比 {stats['pass_rate']}%，样本不均衡，结果相关度可能不准确。")
    
    result_valid = sum(1 for r in records if r['resultNorm'])
    if stats['valid'] > 0 and result_valid / stats['valid'] < 0.8:
        warnings.append(f"面试结果归一化率 {round(result_valid/stats['valid']*100,1)}%，低于 80%。")
    
    grade_valid = sum(1 for r in records if r['gradeNorm'])
    if stats['valid'] > 0 and grade_valid / stats['valid'] < 0.8:
        warnings.append(f"评级归一化率 {round(grade_valid/stats['valid']*100,1)}%，低于 80%。")
    
    if stats['level_dist'].get('无职级', 0) == stats['valid'] and '社招' in str(stats.get('recruit_dist', {})):
        warnings.append("社招数据没有职级信息，将无法输出分职级段的能力模型。")
    
    # 职级段分布
    for seg in ['10-11级', '7-9级']:
        n = stats['level_dist'].get(seg, 0)
        if 0 < n < 10:
            warnings.append(f"{seg} 仅 {n} 条数据，该职级段模型可能不够准确。")
    
    if len(stats['position_dist']) > 1:
        pos_list = ', '.join(f"{k}({v}条)" for k, v in sorted(stats['position_dist'].items(), key=lambda x: -x[1]))
        warnings.append(f"数据包含 {len(stats['position_dist'])} 个职位：{pos_list}。建模一次只分析一个职位。")
    
    # BG 数量 > 1
    if len(stats['bg_dist']) > 1:
        bg_list = ', '.join(f"{k}({v}条)" for k, v in sorted(stats['bg_dist'].items(), key=lambda x: -x[1]))
        warnings.append(f"数据包含 {len(stats['bg_dist'])} 个 BG：{bg_list}。请确认数据来源是否正确。")
    
    # 同时包含社招和校招
    if len(stats['recruit_dist']) > 1:
        rc_list = ', '.join(f"{k}({v}条)" for k, v in stats['recruit_dist'].items())
        warnings.append(f"数据同时包含{rc_list}，将按招聘类型分别生成报告。")
    
    # dept 100% 缺失
    if stats['valid'] > 0 and not stats['dept_dist']:
        warnings.append("部门字段全部缺失，将无法进行部门差异分析。")
    
    return blockers, warnings

def generate_report(stats, blockers, warnings):
    """生成 Markdown 质量报告"""
    lines = ['## 数据清洗报告', '']
    
    # 总体
    lines.append('### 总体概况')
    lines.append(f'- **原始记录数**: {stats["total_raw"]} 条')
    lines.append(f'- **有效记录数**: {stats["valid"]} 条')
    filter_parts = []
    if stats['empty_rows'] > 0:
        filter_parts.append(f"空行 {stats['empty_rows']}")
    if stats['short_comment'] > 0:
        filter_parts.append(f"无评价/过短 {stats['short_comment']}")
    if stats['level_filtered'] > 0:
        filter_parts.append(f"非7-11级 {stats['level_filtered']}")
    if stats['has_huoshui'] > 0:
        filter_parts.append(f"活水数据 {stats['has_huoshui']}")
    if filter_parts:
        lines.append(f'- **过滤明细**: {", ".join(filter_parts)}')
    lines.append('')
    
    # 质量检查
    if blockers:
        lines.append('### ❌ 阻断性问题')
        for b in blockers:
            lines.append(f'- {b}')
        lines.append('')
    
    if warnings:
        lines.append('### ⚠️ 警告')
        for w in warnings:
            lines.append(f'- {w}')
        lines.append('')
    
    if not blockers:
        lines.append('### ✅ 质量检查通过')
        lines.append('')
    
    # 字段映射
    lines.append('### 字段映射')
    lines.append('| 标准字段 | 映射来源 |')
    lines.append('|---------|---------|')
    for field in FIELD_PRIORITY[:9]:
        source = stats['field_map'].get(field, '未映射')
        icon = '✅' if source != '未映射' else '❌'
        lines.append(f'| {field} | {icon} {source} |')
    lines.append('')
    
    # 按部门×招聘类型分布
    lines.append('### 按部门×招聘类型分布')
    lines.append('| 部门 | 招聘类型 | 记录数 | 通过样本占比 | 评级分布 |')
    lines.append('|------|---------|--------|------------|---------|')
    cross = stats.get('cross_dept_recruit', {})
    if cross:
        # 按记录数降序排序
        sorted_cross = sorted(cross.items(), key=lambda x: -x[1]['count'])
        for key, info in sorted_cross:
            dept, rc = key.split('|', 1)
            pass_pct = round(info['pass_count'] / max(1, info['count']) * 100, 1)
            gd = ' / '.join(f"{g}({n})" for g, n in sorted(info['grade_dist'].items()))
            lines.append(f'| {dept} | {rc} | {info["count"]} 条 | {pass_pct}% | {gd} |')
    else:
        for dept, count in stats['dept_dist'].items():
            lines.append(f'| {dept} | — | {count} 条 | — | — |')
    lines.append('')
    
    # 评级分布
    grade_str = ' / '.join(f"{g}({n})" for g, n in sorted(stats['grade_dist'].items()))
    lines.append(f'- **评级分布**: {grade_str}')
    lines.append(f'- **通过样本占比**: {stats["pass_rate"]}%')
    
    # 职级段
    level_str = ' / '.join(f"{k}({v})" for k, v in stats['level_dist'].items() if v > 0)
    lines.append(f'- **职级段**: {level_str}')
    lines.append('')
    
    return '\n'.join(lines)


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='面试数据清洗（Skill1 一键脚本）')
    parser.add_argument('--input', required=True, help='原始数据文件路径（Excel/CSV）')
    parser.add_argument('--output', default='cleaned_data.json', help='输出 JSON 文件路径')
    parser.add_argument('--report', default='quality_report.md', help='质量报告输出路径')
    args = parser.parse_args()
    
    if not os.path.exists(args.input):
        print(f'❌ 文件不存在: {args.input}', file=sys.stderr)
        sys.exit(1)
    
    print(f'处理文件: {args.input}')
    
    # 1. 处理
    records, stats = process_file(args.input)
    
    # 2. 质量检查
    blockers, warnings = quality_check(records, stats)
    
    # 3. 生成报告
    report = generate_report(stats, blockers, warnings)
    with open(args.report, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f'质量报告: {args.report}')
    
    # 4. 输出 JSON
    output = {
        'records': records,
        'metadata': {
            'totalRaw': stats['total_raw'],
            'validCount': stats['valid'],
            'passCount': stats['pass_count'],
            'failCount': stats['fail_count'],
            'passRate': stats['pass_rate'],
            'positionDist': stats['position_dist'],
            'bgDist': stats['bg_dist'],
            'deptDist': stats['dept_dist'],
            'gradeDist': stats['grade_dist'],
            'recruitDist': stats['recruit_dist'],
            'levelDist': stats['level_dist'],
            'fieldMap': stats['field_map'],
            'importTime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    }
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f'标准化数据: {args.output}')
    
    # 5. 控制台摘要
    print(f'\n=== 清洗结果 ===')
    print(f'  原始: {stats["total_raw"]} 条')
    print(f'  有效: {stats["valid"]} 条')
    print(f'  通过样本占比: {stats["pass_rate"]}%')
    print(f'  评级: {" / ".join(f"{g}({n})" for g, n in sorted(stats["grade_dist"].items()))}')
    
    if blockers:
        print(f'\n❌ 阻断性问题:')
        for b in blockers:
            print(f'  - {b}')
        print(f'\n数据不满足建模要求，请修复后重试。')
        sys.exit(1)
    
    if warnings:
        print(f'\n⚠️ 警告:')
        for w in warnings:
            print(f'  - {w}')
    
    print(f'\n✅ 清洗完成，可继续建模。')


if __name__ == '__main__':
    main()
