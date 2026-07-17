from __future__ import annotations

import csv
import datetime as dt
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency for xlsx header fallback
    load_workbook = None

TEXT_FIELDS = [
    "集体面试评语",
    "初试评语",
    "复试评语",
    "GM/EVP面试评语",
    "HR资格面试评语",
    "实习经历",
]

# 基于2022-2026补标数据（16,906条，整体毁约率11.3%），样本>=20，毁约率显著高于均值（delta > +3.5pp）的学校
HIGH_BREACH_RATE_SCHOOLS = {
    "Duke University", "西南财经大学", "Northeastern University", "University of Pennsylvania",
    "中央财经大学", "中国人民大学", "中国科学技术大学", "北京交通大学", "广东外语外贸大学",
    "对外经济贸易大学", "上海外国语大学", "中南财经政法大学", "西安交通大学", "吉林动画学院",
    "北京大学", "东南大学", "University of Southern California", "西北工业大学", "哈尔滨工程大学",
    "国防科技大学", "Cornell University", "南洋理工大学", "苏州大学", "中国科学院大学",
    "清华大学", "北京航空航天大学", "北京理工大学",
}

# 基于同批数据，样本>=20，毁约率显著低于均值（delta < -3.5pp）的学校
LOW_BREACH_RATE_SCHOOLS = {
    "北京师范大学", "上海大学", "天津大学", "武汉理工大学", "香港大学", "Columbia University",
    "广州大学", "香港中文大学", "University College London", "北京科技大学", "福州大学",
    "暨南大学", "深圳大学", "广东工业大学", "杭州电子科技大学", "深圳技术大学", "东华大学",
    "长安大学", "伦敦大学学院", "中国地质大学", "西北大学", "北京电影学院",
    "The Hong Kong University of Science and Technology", "香港理工大学", "墨尔本大学",
    "华中师范大学", "香港科技大学", "中央美术学院", "香港城市大学", "爱丁堡大学",
    "华南农业大学", "合肥工业大学", "哥伦比亚大学", "悉尼大学", "武汉理工大学",
}

BIG_COMPANIES = [
    "字节", "阿里", "美团", "华为", "百度", "京东", "快手", "小红书", "拼多多", "米哈游", "英伟达",
    "滴滴", "网易", "携程", "Amazon", "Google", "Microsoft", "Meta", "OpenAI",
]

EXTERNAL_SIGNAL_WORDS = [
    "其他offer", "保底", "在等", "第3轮", "第4轮", "其他公司面试", "其它公司面试", "并行", "同步在申请",
    "保底offer", "offer比较", "多offer",
]

MISMATCH_RISK_WORDS = [
    "不熟悉", "不接近", "没有工程实践", "更感兴趣", "转方向", "偏研究", "偏学术", "继续考察",
    "岗位不对口", "离家太远", "异地", "回国工作", "不考虑在", "不考虑长期", "只考虑", "更倾向", "更想要",
]

STABILITY_RISK_WORDS = [
    "读博", "读博机会", "保研", "继续深造", "深造", "出国", "留学", "回国计划", "优先去读博",
    "考公", "考编", "国企", "银行系统", "体制内", "年底回国", "影响入职", "也在申请",
    "已入职", "社保", "转正答辩", "留用流程", "正式offer",
]

PROTECTIVE_WORDS = [
    "就业优先腾讯", "优先腾讯", "首选腾讯", "只考虑腾讯", "更喜欢腾讯", "腾讯意向很高", "腾讯优先",
    "加入意愿清晰", "支持留用", "长期发展", "稳定性较高", "明确意向",
]

STRONG_TENCENT_WORDS = ["对腾讯意愿度最高", "优先考虑腾讯的留用", "优先考虑腾讯", "首选腾讯", "腾讯优先"]
CURRENT_COMPETITOR_STATUS_WORDS = ["转正", "留用", "已入职", "社保", "转正答辩", "留用流程", "正式offer"]
STABILITY_NEGATIVE_PATTERNS = [
    r"不考虑继续深造", r"无读研深造计划", r"期望直接就业", r"直接就业", r"不考虑读博", r"毕业后到工业界",
    r"达不到保研要求", r"不考虑留用", r"不继续留用", r"没有选择留用",
]
STABILITY_HISTORY_PATTERNS = [
    r"原本.*出国", r"原计划.*出国", r"因为疫情原因没有出国", r"疫情原因没有出国", r"本科毕业原计划出国", r"保研本校",
]
CITY_WORDS = ["深圳", "广州", "北京", "上海", "杭州", "南京", "成都", "武汉"]

LEAKY_OR_POST_EVENT_FIELDS = {
    "签约状态", "毁约/拒签原因", "拒签时间", "毁约时间", "拒签原因", "拒签原因备注", "放弃类型", "已签时间", "最新员工状态",
}


@dataclass
class ScoreResult:
    attention_score: int
    attention_level: str
    focus_flag: str
    priority: str
    stability_score: int
    stability_level: str
    stable_flag: str
    dimension: str
    second_dimension: str
    dimensions: dict[str, int]
    reasons: list[str]
    protective_signals: list[str]
    summary: str
    action: str
    owner: str


@dataclass
class ProcessFeatures:
    hr_to_salary_days: int | None
    salary_to_approve_days: int | None
    approve_to_send_days: int | None
    offer_to_resume_update_days: int | None
    offer_to_status_update_days: int | None


def norm(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in {"", "-", "—", "None", "nan", "NaN"} else s


def text_join(row: dict, fields: Iterable[str]) -> str:
    return "\n".join(norm(row.get(f, "")) for f in fields if norm(row.get(f, "")))


FIELD_ALIASES = {
    "简历ID": ["简历ID", "resume_id"],
    "招聘年份": ["招聘年份", "recruit_year"],
    "最高学历学校": ["最高学历学校", "highest_school"],
    "最高学历": ["最高学历", "highest_degree"],
    "最高学历专业": ["最高学历专业", "highest_speciality"],
    "职位": ["职位", "position_name_cn"],
    "工作地": ["工作地", "w_city"],
    "实习经历": ["实习经历", "practice_exp"],
    "实习公司": ["实习公司", "employer_names"],
    "OFFER类型": ["OFFER类型", "offer_type", "offer_type_name"],
    "招聘渠道": ["招聘渠道", "recruit_channel"],
    "是否当前最新流程": ["是否当前最新流程", "lastest_flow_flag_name"],
    "HR面试通过时间": ["HR面试通过时间", "hr_pass_time"],
    "HR定薪提交时间": ["HR定薪提交时间", "hr_salary_submit_time"],
    "OFFER审批通过时间": ["OFFER审批通过时间", "offer_approve_time"],
    "发送offer时间": ["发送offer时间", "offer_send_time"],
    "简历更新时间": ["简历更新时间", "resume_update_time"],
    "简历状态更新时间": ["简历状态更新时间", "resume_status_update_time"],
    "集体面试评语": ["集体面试评语", "group_interview_comment"],
    "初试评语": ["初试评语", "first_interview_comment"],
    "复试评语": ["复试评语", "second_interview_comment"],
    "GM/EVP面试评语": ["GM/EVP面试评语", "gm_interview_comment"],
    "HR资格面试评语": ["HR资格面试评语", "hr_interview_comment"],
}


def load_xlsx_header(csv_path: Path) -> list[str] | None:
    if load_workbook is None:
        return None
    xlsx_candidates = [csv_path.with_suffix(".xlsx"), csv_path.with_suffix(".XLSX")]
    for xlsx in xlsx_candidates:
        if xlsx.exists():
            wb = load_workbook(xlsx, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header = list(next(ws.iter_rows(values_only=True)))
            return [norm(h) for h in header]
    return None



def normalize_row_keys(row: dict) -> dict:
    normalized = dict(row)
    for target, aliases in FIELD_ALIASES.items():
        if norm(normalized.get(target, "")):
            continue
        for alias in aliases:
            value = normalized.get(alias, "")
            if norm(value):
                normalized[target] = value
                break
    return normalized



def read_csv_rows(csv_path: str | Path) -> tuple[list[dict], str]:
    csv_path = Path(csv_path)
    header_fallback = load_xlsx_header(csv_path)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        raw_header = next(reader)
        rows = list(reader)
    if header_fallback and len(header_fallback) == len(raw_header):
        header = header_fallback
        encoding = "utf-8-sig + xlsx_header_fallback"
    else:
        header = raw_header
        encoding = "utf-8-sig"
    dict_rows: list[dict] = []
    for row in rows:
        padded = row + [""] * (len(header) - len(row))
        dict_rows.append(normalize_row_keys({header[i]: padded[i] for i in range(len(header))}))
    return dict_rows, encoding



def read_json_rows(json_path: str | Path) -> tuple[list[dict], str]:
    json_path = Path(json_path)
    data = json.loads(json_path.read_text(encoding="utf-8-sig"))
    if isinstance(data, dict):
        if isinstance(data.get("records"), list):
            data = data["records"]
        elif isinstance(data.get("data"), list):
            data = data["data"]
        else:
            data = [data]
    if not isinstance(data, list):
        raise ValueError("JSON 输入必须是 list[dict] 或包含 records/data 的对象")
    return [normalize_row_keys(row) for row in data if isinstance(row, dict)], "json"



def read_input_rows(input_path: str | Path) -> tuple[list[dict], str]:
    input_path = Path(input_path)
    if input_path.suffix.lower() == ".json":
        return read_json_rows(input_path)
    return read_csv_rows(input_path)


def count_company_hits(text: str) -> tuple[int, list[str]]:
    hits = [c for c in BIG_COMPANIES if c.lower() in text.lower()]
    return len(hits), hits


def count_keyword_hits(text: str, keywords: list[str]) -> tuple[int, list[str]]:
    hits = [kw for kw in keywords if kw.lower() in text.lower()]
    return len(hits), hits


def internship_company_count(row: dict) -> int:
    raw = norm(row.get("实习公司", ""))
    if not raw:
        return 0
    return len([x.strip() for x in raw.split("##") if norm(x)])


def internship_company_hits(row: dict) -> list[str]:
    raw = norm(row.get("实习公司", ""))
    if not raw:
        return []
    return [c for c in BIG_COMPANIES if c.lower() in raw.lower()]


def city_in_text(text: str) -> list[str]:
    return [c for c in CITY_WORDS if c in text]


# 职位匹配函数已移除。
# 分析结论（2026-06-08）：跨大类投递vs录用不一致的毁约率11.2% ≈ 整体均值11.3%，
# 精确匹配毁约率10.4%，差异不到1个百分点。职位名称不一致主要源于校招内外部命名体系差异，
# 不宜作为风险或保护信号。


def base_city_label(text: str) -> str:
    for city in CITY_WORDS:
        if city in norm(text):
            return city
    return ""


def city_compatible(preferred_city: str, work_city: str) -> bool:
    preferred_city = norm(preferred_city)
    work_city = norm(work_city)
    if not preferred_city or not work_city:
        return False
    if preferred_city in work_city or work_city in preferred_city:
        return True
    if preferred_city == "广深" and any(c in work_city for c in ["广州", "深圳"]):
        return True
    if work_city == "广深" and any(c in preferred_city for c in ["广州", "深圳"]):
        return True
    if preferred_city in {"广州", "深圳"} and work_city in {"广州", "深圳"}:
        return True
    return False


def has_explicit_city_conflict(text: str, work_city: str) -> bool:
    text = norm(text)
    work_base_city = base_city_label(work_city)
    if not text or not work_base_city:
        return False
    if "广深" in text and work_base_city in {"广州", "深圳"}:
        return False
    preferred_patterns = ["只考虑", "更倾向", "倾向", "希望", "优先", "想去", "工作地", "工作城市", "期望城市"]
    for marker in preferred_patterns:
        if re.search(rf"{marker}.{{0,10}}{work_base_city}", text) or re.search(rf"{work_base_city}.{{0,10}}{marker}", text):
            return False
    for city in city_in_text(text):
        if city_compatible(city, work_base_city):
            continue
        patterns = [
            rf"(只考虑|更倾向|倾向|希望|优先|想去|工作地|工作城市|期望城市).{{0,8}}{city}",
            rf"{city}.{{0,8}}(只考虑|更倾向|倾向|希望|优先|想去|工作地|工作城市|期望城市)",
        ]
        if any(re.search(p, text) for p in patterns):
            return True
    return False


def matched_patterns(text: str, patterns: list[str]) -> list[str]:
    return [pattern for pattern in patterns if re.search(pattern, text)]


def sorted_dimensions(dims: dict[str, int]) -> list[tuple[str, int]]:
    return sorted(dims.items(), key=lambda item: (-item[1], item[0]))


def parse_date(value: str) -> dt.datetime | None:
    value = norm(value)
    if not value:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
    ):
        try:
            return dt.datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def days_between(start: str, end: str) -> int | None:
    left = parse_date(start)
    right = parse_date(end)
    if not left or not right:
        return None
    return (right - left).days


def build_process_features(row: dict) -> ProcessFeatures:
    return ProcessFeatures(
        hr_to_salary_days=days_between(row.get("HR面试通过时间", ""), row.get("HR定薪提交时间", "")),
        salary_to_approve_days=days_between(row.get("HR定薪提交时间", ""), row.get("OFFER审批通过时间", "")),
        approve_to_send_days=days_between(row.get("OFFER审批通过时间", ""), row.get("发送offer时间", "")),
        offer_to_resume_update_days=days_between(row.get("发送offer时间", ""), row.get("简历更新时间", "")),
        offer_to_status_update_days=days_between(row.get("发送offer时间", ""), row.get("简历状态更新时间", "")),
    )


def score_text_dimensions(row: dict) -> tuple[dict[str, int], list[str], list[str]]:
    dims = {"竞争变动": 0, "匹配偏差": 0, "去向不稳": 0}
    reasons: list[str] = []
    protections: list[str] = []
    all_text = text_join(row, TEXT_FIELDS)
    work_city = norm(row.get("工作地", ""))
    study_city = norm(row.get("目前就读地", ""))

    company_hit_n, company_hits = count_company_hits(all_text)
    ext_kw_n, ext_kw_hits = count_keyword_hits(all_text, EXTERNAL_SIGNAL_WORDS)
    mismatch_kw_n, mismatch_hits = count_keyword_hits(all_text, MISMATCH_RISK_WORDS)
    stability_kw_n, stability_hits = count_keyword_hits(all_text, STABILITY_RISK_WORDS)
    protect_kw_n, protect_hits = count_keyword_hits(all_text, PROTECTIVE_WORDS)
    strong_tencent_n, strong_tencent_hits = count_keyword_hits(all_text, STRONG_TENCENT_WORDS)
    stability_negative_hits = matched_patterns(all_text, STABILITY_NEGATIVE_PATTERNS)
    stability_history_hits = matched_patterns(all_text, STABILITY_HISTORY_PATTERNS)

    if company_hit_n:
        gain = min(2 + company_hit_n * 2, 8)
        dims["竞争变动"] += gain
        reasons.append(f"文本出现外部大厂/强竞争信号：{', '.join(company_hits[:4])}")
    if ext_kw_n:
        gain = min(2 + ext_kw_n * 2, 6)
        dims["竞争变动"] += gain
        reasons.append(f"文本出现并行流程/offer比较信号：{', '.join(ext_kw_hits[:4])}")

    if mismatch_kw_n:
        gain = min(5 + mismatch_kw_n * 2, 10)
        dims["匹配偏差"] += gain
        reasons.append(f"文本出现岗位/城市/方向偏差信号：{', '.join(mismatch_hits[:4])}")

    if study_city and work_city and study_city not in work_city and all(c not in work_city for c in city_in_text(study_city)):
        dims["匹配偏差"] += 3
        reasons.append(f"目前就读地“{study_city}”与工作地“{work_city}”不一致")

    if has_explicit_city_conflict(all_text, work_city):
        dims["匹配偏差"] += 4
        reasons.append(f"文本表达的城市偏好与工作地“{work_city}”存在差异")

    if stability_kw_n:
        gain = min(5 + stability_kw_n * 3, 11)
        dims["去向不稳"] += gain
        reasons.append(f"文本出现深造/体制内/留用去向信号：{', '.join(stability_hits[:4])}")

    competitor_status_n, competitor_status_hits = count_keyword_hits(all_text, CURRENT_COMPETITOR_STATUS_WORDS)
    if any(phrase in all_text for phrase in ["不考虑留用", "不继续留用", "没有选择留用"]):
        competitor_status_n = 0
        competitor_status_hits = []
    if competitor_status_n and company_hit_n:
        dims["去向不稳"] += min(2 + competitor_status_n, 4)
        reasons.append(f"当前处于外部公司留用/转正/已入职阶段：{', '.join(competitor_status_hits[:4])}")

    if "至今" in norm(row.get("实习经历", "")) and company_hit_n:
        dims["去向不稳"] += 2
        reasons.append("当前仍在外部强竞争公司/机构经历中")

    if stability_negative_hits:
        dims["去向不稳"] = max(0, dims["去向不稳"] - 6)
        protections.append("文本出现明确的直接就业/不继续深造信号")
    if stability_history_hits:
        dims["去向不稳"] = max(0, dims["去向不稳"] - 4)
        protections.append("文本中的深造/出国信息更偏历史背景")
    if strong_tencent_n:
        dims["竞争变动"] = max(0, dims["竞争变动"] - 2)
        dims["去向不稳"] = max(0, dims["去向不稳"] - 3)
        protections.append(f"存在较强的腾讯优先承诺：{', '.join(strong_tencent_hits[:3])}")
    if protect_kw_n:
        reduce_n = min(1 + protect_kw_n, 3)
        dims["匹配偏差"] = max(0, dims["匹配偏差"] - reduce_n)
        dims["去向不稳"] = max(0, dims["去向不稳"] - reduce_n)
        protections.append(f"存在明确的腾讯优先/加入意愿信号：{', '.join(protect_hits[:4])}")

    dims["竞争变动"] = min(max(dims["竞争变动"], 0), 12)
    dims["匹配偏差"] = min(max(dims["匹配偏差"], 0), 18)
    dims["去向不稳"] = min(max(dims["去向不稳"], 0), 18)
    return dims, reasons, protections


def score_time_process_dimension(row: dict, pf: ProcessFeatures) -> tuple[int, list[str], list[str]]:
    score = 0
    reasons: list[str] = []
    protections: list[str] = []

    if pf.hr_to_salary_days is not None:
        if pf.hr_to_salary_days > 30:
            score += 6
            reasons.append(f"HR面试通过到定薪提交耗时{pf.hr_to_salary_days}天，流程推进偏慢")
        elif pf.hr_to_salary_days > 14:
            score += 3
            reasons.append(f"HR面试通过到定薪提交耗时{pf.hr_to_salary_days}天，存在流程拖延")
        elif 0 <= pf.hr_to_salary_days <= 7:
            protections.append(f"HR面试通过到定薪提交仅{pf.hr_to_salary_days}天，流程推进较顺")

    if pf.salary_to_approve_days is not None:
        if pf.salary_to_approve_days > 30:
            score += 5
            reasons.append(f"定薪提交到offer审批通过耗时{pf.salary_to_approve_days}天，审批链路偏长")
        elif pf.salary_to_approve_days > 7:
            score += 2
            reasons.append(f"定薪提交到offer审批通过耗时{pf.salary_to_approve_days}天")
        elif 0 <= pf.salary_to_approve_days <= 3:
            protections.append(f"定薪提交到offer审批通过仅{pf.salary_to_approve_days}天，审批较快")

    if pf.approve_to_send_days is not None:
        if pf.approve_to_send_days > 7:
            score += 4
            reasons.append(f"offer审批通过后{pf.approve_to_send_days}天才发放offer")
        elif pf.approve_to_send_days > 3:
            score += 1
            reasons.append(f"offer审批通过后{pf.approve_to_send_days}天发放，节奏一般")
        elif 0 <= pf.approve_to_send_days <= 3:
            protections.append(f"offer审批通过后{pf.approve_to_send_days}天内发放，承接较快")

    if pf.offer_to_resume_update_days is not None and pf.offer_to_resume_update_days > 7:
        score += 3
        reasons.append(f"发放offer后{pf.offer_to_resume_update_days}天简历仍有更新，可能仍在持续比较机会")
    if pf.offer_to_status_update_days is not None and pf.offer_to_status_update_days > 7:
        score += 2
        reasons.append(f"发放offer后{pf.offer_to_status_update_days}天简历状态仍有变化")

    return min(score, 18), reasons, protections


def score_structure_dimension(row: dict) -> tuple[int, list[str], int, list[str]]:
    score = 0
    reasons: list[str] = []
    protect_points = 0
    protections: list[str] = []

    school = norm(row.get("最高学历学校", ""))
    degree = norm(row.get("最高学历", ""))
    offer_type = norm(row.get("OFFER类型", ""))
    offer_pos = norm(row.get("职位", ""))
    hr_grade = norm(row.get("HR资格面试等级", ""))
    gm_grade = norm(row.get("GM/EVP面试等级", ""))
    overall_assess = norm(row.get("整体测评状态", ""))
    internship_type = norm(row.get("实习类型", ""))
    raw_intern_company = norm(row.get("实习公司", ""))
    channel = norm(row.get("招聘渠道", ""))
    latest_flag = norm(row.get("是否当前最新流程", ""))

    intern_company_n = internship_company_count(row)
    intern_big_hits = internship_company_hits(row)
    has_tencent_intern = "腾讯" in raw_intern_company

    if school in HIGH_BREACH_RATE_SCHOOLS and degree in {"硕士研究生", "博士研究生"}:
        score += 2
        reasons.append(f"{school}历史毁约率偏高，结构性参考")
    if school in LOW_BREACH_RATE_SCHOOLS:
        protect_points += 1
        protections.append(f"{school}历史毁约率低于均值")
    if degree == "博士研究生" and any(x in offer_pos for x in ["研究", "算法", "应用研究"]):
        score += 2
        reasons.append("博士研究生 + 研究/算法岗位，外部可选项通常更多")
    if offer_type in {"S档offer", "A+档offer", "特殊档"}:
        score += 2
        reasons.append(f"{offer_type}，通常对应更强的市场竞争环境")
    if intern_company_n >= 3 and len(intern_big_hits) >= 2:
        score += 3
        reasons.append(f"拥有{intern_company_n}段实习且含多家头部公司经历，外部选择面较宽")
    elif len(intern_big_hits) >= 2:
        score += 2
        reasons.append(f"存在多家头部公司实习背景：{', '.join(intern_big_hits[:4])}")

    if hr_grade in {"A+", "S"}:
        protect_points += 3
        protections.append(f"HR资格面试等级为{hr_grade}")
    if gm_grade in {"A+", "S"}:
        protect_points += 2
        protections.append(f"GM/EVP面试等级为{gm_grade}")
    if "绿灯" in overall_assess:
        protect_points += 4
        protections.append("整体测评状态为绿灯")
    if has_tencent_intern:
        protect_points += 2
        protections.append("存在腾讯实习/留用背景")
    if internship_type == "留用":
        protect_points += 2
        protections.append("实习类型为留用")
    if channel == "自行投递":
        protect_points += 1
        protections.append("自主投递意愿较明确")
    if latest_flag == "是":
        protect_points += 1
        protections.append("当前记录为最新流程")

    return min(score, 14), reasons, protect_points, protections


def derive_priority(attention_score: int, dims: dict[str, int], stability_score: int) -> tuple[str, str, str]:
    time_process = dims.get("时间流程", 0)
    mismatch = dims.get("匹配偏差", 0)
    unstable = dims.get("去向不稳", 0)
    if attention_score >= 42 or (attention_score >= 36 and (time_process >= 8 or mismatch >= 9 or unstable >= 9)):
        priority = "P1"
    elif attention_score >= 30 or (attention_score >= 24 and time_process >= 8):
        priority = "P2"
    elif attention_score >= 18:
        priority = "P3"
    else:
        priority = "P4"

    if stability_score >= 55 and priority == "P3" and time_process <= 5 and mismatch <= 7 and unstable <= 7:
        priority = "P4"
    elif stability_score >= 50 and priority == "P2" and time_process <= 4 and mismatch <= 6 and unstable <= 6:
        priority = "P3"

    focus_flag = "是" if priority in {"P1", "P2"} else "否"
    level_map = {"P1": "高", "P2": "中高", "P3": "中", "P4": "低"}
    return priority, focus_flag, level_map[priority]


def derive_stability_level(stability_score: int) -> tuple[str, str]:
    if stability_score >= 55:
        return "高", "是"
    if stability_score >= 40:
        return "中", "否（相对稳定）"
    return "低", "否"


def compact(items: list[str], limit: int = 4) -> list[str]:
    dedup: list[str] = []
    seen = set()
    for item in items:
        if item and item not in seen:
            dedup.append(item)
            seen.add(item)
    return dedup[:limit]


def build_summary_action(priority: str, attention_level: str, stability_level: str, stable_flag: str, main_dimension: str, second_dimension: str) -> tuple[str, str, str]:
    if priority == "P1":
        owner = "招聘 + 业务主管 + 导师/面试官"
        summary = f"当前属于{priority}级重点关注，主因集中在{main_dimension}。"
    elif priority == "P2":
        owner = "招聘 + 用人主管"
        summary = f"当前属于{priority}级重点关注，建议围绕{main_dimension}优先核实。"
    elif stable_flag == "是":
        owner = "招聘"
        summary = "当前更接近稳定签约画像，可按常规节奏保温。"
    else:
        owner = "招聘"
        summary = f"当前整体关注等级为{attention_level}，建议结合{main_dimension}做常规核验。"

    if main_dimension == "时间流程":
        action = "优先检查定薪、审批、offer发放与候选人反馈之间是否存在拖延或断点，必要时压缩等待窗口。"
    elif main_dimension == "匹配偏差":
        action = "重点核实岗位内容、团队归属、工作城市与候选人真实预期是否一致，避免因理解偏差流失。"
    elif main_dimension == "去向不稳":
        action = "优先确认深造、国企/考公、外部留用等替代去向是否仍在持续，并明确决策窗口。"
    elif main_dimension == "结构信号":
        action = "结合学历、实习、offer档位等结构信号判断其外部选择面，提前准备差异化卖点。"
    else:
        action = "跟进并行流程进展，突出岗位与团队价值，保持节奏稳定。"

    if stable_flag == "是":
        action += " 同时该同学具备较多稳定签约信号，可适度降低升级干预频率。"
    elif stability_level == "中" and priority not in {"P1", "P2"}:
        action += " 如无新增异常，可维持轻量保温。"

    if second_dimension:
        summary += f" 次要关注点为{second_dimension}。"
    return summary, action, owner


def score_row(row: dict) -> ScoreResult:
    text_dims, text_reasons, text_protections = score_text_dimensions(row)
    process_features = build_process_features(row)
    process_score, process_reasons, process_protections = score_time_process_dimension(row, process_features)
    structure_score, structure_reasons, protect_points, structure_protections = score_structure_dimension(row)

    dims = {
        "竞争变动": text_dims["竞争变动"],
        "匹配偏差": text_dims["匹配偏差"],
        "去向不稳": text_dims["去向不稳"],
        "时间流程": process_score,
        "结构信号": structure_score,
    }
    reasons = compact(text_reasons + process_reasons + structure_reasons, 5)
    protective_signals = compact(text_protections + process_protections + structure_protections, 5)

    attention_score = min(sum(dims.values()), 100)
    stability_score = max(0, min(100, 28 + protect_points * 4 - attention_score))
    stability_level, stable_flag = derive_stability_level(stability_score)
    priority, focus_flag, attention_level = derive_priority(attention_score, dims, stability_score)

    ranked_dims = sorted_dimensions(dims)
    main_dimension = ranked_dims[0][0]
    second_dimension = ranked_dims[1][0] if len(ranked_dims) > 1 and ranked_dims[1][1] > 0 else ""
    summary, action, owner = build_summary_action(priority, attention_level, stability_level, stable_flag, main_dimension, second_dimension)

    return ScoreResult(
        attention_score=attention_score,
        attention_level=attention_level,
        focus_flag=focus_flag,
        priority=priority,
        stability_score=stability_score,
        stability_level=stability_level,
        stable_flag=stable_flag,
        dimension=main_dimension,
        second_dimension=second_dimension,
        dimensions=dims,
        reasons=reasons,
        protective_signals=protective_signals,
        summary=summary,
        action=action,
        owner=owner,
    )


def make_full_row(row: dict, res: ScoreResult) -> dict:
    new_row = dict(row)
    new_row["关注分"] = res.attention_score
    new_row["关注等级"] = res.attention_level
    new_row["是否建议重点关注"] = res.focus_flag
    new_row["关注优先级"] = res.priority
    new_row["稳定签约分"] = res.stability_score
    new_row["稳定签约等级"] = res.stability_level
    new_row["是否识别为相对稳定签约"] = res.stable_flag
    new_row["主关注维度"] = res.dimension
    new_row["次关注维度"] = res.second_dimension
    new_row["关注维度得分"] = json.dumps(res.dimensions, ensure_ascii=False)
    new_row["保护性信号"] = "；".join(res.protective_signals)
    new_row["关注理由"] = "；".join(res.reasons)
    new_row["招聘侧判断摘要"] = res.summary
    new_row["推荐跟进行动"] = res.action
    new_row["推荐介入角色"] = res.owner
    new_row["模型定位说明"] = "该版本用于输出关注建议与稳定签约识别，不直接等同于毁约概率。脚本已排除签约状态、已签/毁约/拒签时间等结果字段。"
    new_row["agent_payload"] = json.dumps(
        {
            "candidate_id": norm(row.get("简历ID", "")),
            "attention_score": res.attention_score,
            "attention_level": res.attention_level,
            "focus": res.focus_flag,
            "priority": res.priority,
            "stability_score": res.stability_score,
            "stability_level": res.stability_level,
            "stable_flag": res.stable_flag,
            "main_dimension": res.dimension,
            "second_dimension": res.second_dimension,
            "dimension_scores": res.dimensions,
            "summary": res.summary,
            "action": res.action,
            "owner": res.owner,
            "reasons": res.reasons,
            "protective_signals": res.protective_signals,
        },
        ensure_ascii=False,
    )
    return new_row


def make_focus_row(row: dict, res: ScoreResult) -> dict:
    return {
        "简历ID": norm(row.get("简历ID", "")),
        "招聘年份": norm(row.get("招聘年份", "")),
        "最高学历学校": norm(row.get("最高学历学校", "")),
        "最高学历": norm(row.get("最高学历", "")),
        "职位": norm(row.get("职位", "")),
        "工作地": norm(row.get("工作地", "")),
        "OFFER类型": norm(row.get("OFFER类型", "")),
        "关注分": res.attention_score,
        "关注等级": res.attention_level,
        "关注优先级": res.priority,
        "稳定签约分": res.stability_score,
        "稳定签约等级": res.stability_level,
        "主关注维度": res.dimension,
        "次关注维度": res.second_dimension,
        "保护性信号": "；".join(res.protective_signals),
        "关注理由": "；".join(res.reasons),
        "招聘侧判断摘要": res.summary,
        "推荐跟进行动": res.action,
        "推荐介入角色": res.owner,
    }


def make_stable_row(row: dict, res: ScoreResult) -> dict:
    return {
        "简历ID": norm(row.get("简历ID", "")),
        "招聘年份": norm(row.get("招聘年份", "")),
        "最高学历学校": norm(row.get("最高学历学校", "")),
        "最高学历": norm(row.get("最高学历", "")),
        "职位": norm(row.get("职位", "")),
        "工作地": norm(row.get("工作地", "")),
        "OFFER类型": norm(row.get("OFFER类型", "")),
        "稳定签约分": res.stability_score,
        "稳定签约等级": res.stability_level,
        "关注分": res.attention_score,
        "关注等级": res.attention_level,
        "主关注维度": res.dimension,
        "保护性信号": "；".join(res.protective_signals),
        "关注理由": "；".join(res.reasons),
        "招聘侧判断摘要": res.summary,
        "推荐跟进行动": res.action,
    }


def write_csv(path: str | Path, rows: list[dict], fieldnames: list[str] | None = None) -> None:
    path = Path(path)
    if not rows:
        if fieldnames is None:
            fieldnames = []
        with path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
        return
    if fieldnames is None:
        fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def default_related_path(output_csv: str | Path, suffix: str, ext: str) -> Path:
    output_csv = Path(output_csv)
    return output_csv.with_name(f"{output_csv.stem}{suffix}{ext}")


def score_file(
    input_csv: str | Path,
    output_csv: str | Path,
    summary_json: str | Path,
    focus_csv: str | Path | None = None,
    stable_csv: str | Path | None = None,
    agent_json: str | Path | None = None,
) -> dict:
    rows, encoding = read_input_rows(input_csv)
    full_rows: list[dict] = []
    focus_rows: list[dict] = []
    stable_rows: list[dict] = []
    agent_records: list[dict] = []
    priority_counter = Counter()
    attention_counter = Counter()
    stable_counter = Counter()
    dim_counter = Counter()
    attention_scores: list[int] = []
    stability_scores: list[int] = []

    for row in rows:
        res = score_row(row)
        full_row = make_full_row(row, res)
        full_rows.append(full_row)
        attention_scores.append(res.attention_score)
        stability_scores.append(res.stability_score)
        priority_counter[res.priority] += 1
        attention_counter[res.attention_level] += 1
        stable_counter[res.stability_level] += 1
        dim_counter[res.dimension] += 1

        if res.priority in {"P1", "P2"}:
            focus_rows.append(make_focus_row(row, res))
        if res.stable_flag == "是" and res.priority in {"P3", "P4"}:
            stable_rows.append(make_stable_row(row, res))

        agent_records.append(
            {
                "candidate_id": norm(row.get("简历ID", "")),
                "year": norm(row.get("招聘年份", "")),
                "position": norm(row.get("职位", "")),
                "work_city": norm(row.get("工作地", "")),
                "school": norm(row.get("最高学历学校", "")),
                "degree": norm(row.get("最高学历", "")),
                "offer_type": norm(row.get("OFFER类型", "")),
                "attention_score": res.attention_score,
                "attention_level": res.attention_level,
                "priority": res.priority,
                "focus": res.focus_flag,
                "stability_score": res.stability_score,
                "stability_level": res.stability_level,
                "stable_flag": res.stable_flag,
                "main_dimension": res.dimension,
                "second_dimension": res.second_dimension,
                "dimension_scores": res.dimensions,
                "summary": res.summary,
                "recommended_action": res.action,
                "recommended_owner": res.owner,
                "reasons": res.reasons,
                "protective_signals": res.protective_signals,
            }
        )

    output_csv = Path(output_csv)
    if focus_csv is None:
        focus_csv = default_related_path(output_csv, "_重点关注名单", ".csv")
    if stable_csv is None:
        stable_csv = default_related_path(output_csv, "_稳定签约名单", ".csv")
    if agent_json is None:
        agent_json = default_related_path(output_csv, "_agent", ".json")

    focus_rows = sorted(focus_rows, key=lambda row: (-int(row["关注分"]), row["简历ID"]))
    stable_rows = sorted(stable_rows, key=lambda row: (-int(row["稳定签约分"]), int(row["关注分"]), row["简历ID"]))
    agent_records = sorted(agent_records, key=lambda row: (-int(row["attention_score"]), -int(row["stability_score"]), row["candidate_id"]))

    write_csv(output_csv, full_rows)
    write_csv(focus_csv, focus_rows, fieldnames=list(focus_rows[0].keys()) if focus_rows else list(make_focus_row({}, ScoreResult(0, "低", "否", "P4", 0, "低", "否", "竞争变动", "", {"竞争变动": 0, "匹配偏差": 0, "去向不稳": 0, "时间流程": 0, "结构信号": 0}, [], [], "", "", "招聘")).keys()))
    write_csv(stable_csv, stable_rows, fieldnames=list(stable_rows[0].keys()) if stable_rows else list(make_stable_row({}, ScoreResult(0, "低", "否", "P4", 0, "低", "否", "竞争变动", "", {"竞争变动": 0, "匹配偏差": 0, "去向不稳": 0, "时间流程": 0, "结构信号": 0}, [], [], "", "", "招聘")).keys()))

    summary = {
        "input": str(input_csv),
        "encoding": encoding,
        "rows": len(full_rows),
        "model_positioning": "关注建议 + 稳定签约识别，不直接等同于毁约概率",
        "excluded_fields": sorted(LEAKY_OR_POST_EVENT_FIELDS),
        "avg_attention_score": round(sum(attention_scores) / len(attention_scores), 1) if attention_scores else 0,
        "median_attention_score": sorted(attention_scores)[len(attention_scores) // 2] if attention_scores else 0,
        "avg_stability_score": round(sum(stability_scores) / len(stability_scores), 1) if stability_scores else 0,
        "median_stability_score": sorted(stability_scores)[len(stability_scores) // 2] if stability_scores else 0,
        "attention_levels": dict(attention_counter),
        "priorities": dict(priority_counter),
        "stability_levels": dict(stable_counter),
        "top_dimension": dict(dim_counter),
        "focus_count": len(focus_rows),
        "stable_count": len(stable_rows),
        "full_output": str(output_csv),
        "focus_output": str(focus_csv),
        "stable_output": str(stable_csv),
        "agent_output": str(agent_json),
    }
    Path(summary_json).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    agent_payload = {
        "meta": summary,
        "records": agent_records,
    }
    Path(agent_json).write_text(json.dumps(agent_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="V4 关注建议与稳定签约识别脚本（支持 CSV 或 JSON 候选人明细输入）")
    parser.add_argument("input_csv", help="输入候选人明细文件，支持 .csv / .json")
    parser.add_argument("--output-csv", default="V4关注与稳定签约评分结果.csv")
    parser.add_argument("--summary-json", default="V4关注与稳定签约评分摘要.json")
    parser.add_argument("--focus-csv", default=None)
    parser.add_argument("--stable-csv", default=None)
    parser.add_argument("--agent-json", default=None)
    args = parser.parse_args()

    summary = score_file(
        input_csv=args.input_csv,
        output_csv=args.output_csv,
        summary_json=args.summary_json,
        focus_csv=args.focus_csv,
        stable_csv=args.stable_csv,
        agent_json=args.agent_json,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
